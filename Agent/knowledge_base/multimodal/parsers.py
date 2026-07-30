"""将允许的资料转换为项目内部的标准化解析结果。"""

from __future__ import annotations

import importlib.metadata
import importlib.util
import csv
import hashlib
import json
import multiprocessing
import os
import time
from dataclasses import dataclass, field
from io import BytesIO
from collections.abc import Iterator
from pathlib import Path

from .contracts import IngestionIssue, IssueSeverity

TEXT_SUFFIXES = {".txt", ".md", ".markdown"}
TABLE_SUFFIXES = {".csv", ".xlsx"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"}
PDF_SUFFIXES = {".pdf"}
SUPPORTED_SUFFIXES = TEXT_SUFFIXES | TABLE_SUFFIXES | IMAGE_SUFFIXES | PDF_SUFFIXES
PAGE_QUALITY_GATE_VERSION = "page-quality-v1"
PAGE_QUALITY_MIN_TEXT_COVERAGE = 0.25
# ponytail: 固定常量进 manifest；以后要可配再读 env，不要现在抽象成配置系统
TEXT_SPLIT = {"target_chars": 800, "max_chars": 1200, "overlap_chars": 120}
_RAPID_OCR_ENGINE: object | None = None
_OCR_FINGERPRINT: dict[str, str] | None = None
_RAPIDOCR_MODEL_NAMES = (
    "PP-OCRv6_det_small.onnx",
    "ch_ppocr_mobile_v2.0_cls_mobile.onnx",
    "PP-OCRv6_rec_small.onnx",
)


@dataclass(frozen=True)
class OcrResult:
    """本地 OCR 引擎一次推理的分类结果，区分无文字、依赖缺失与推理失败。

    ponytail: 不进入 KnowledgeUnit 公共接口，仅作 parser 内部契约和 manifest 指纹来源。
    """

    text: str
    engine: str
    engine_version: str
    model_fingerprint: str
    line_count: int
    mean_confidence: float
    elapsed_ms: int
    status: str  # "ok" | "no_text" | "failed" | "engine_unavailable"
    failure_type: str = ""


def ocr_fingerprint() -> dict[str, str]:
    """返回 RapidOCR 包与固定本地模型文件的内容指纹，不加载模型。"""
    global _OCR_FINGERPRINT
    if _OCR_FINGERPRINT is not None:
        return dict(_OCR_FINGERPRINT)
    engine = "rapidocr-onnxruntime"
    try:
        version = importlib.metadata.version("rapidocr-onnxruntime")
    except importlib.metadata.PackageNotFoundError:
        version = "unknown"
    spec = importlib.util.find_spec("rapidocr")
    model_root = Path(next(iter(spec.submodule_search_locations), "")) / "models" if spec and spec.submodule_search_locations else None
    model_paths = [model_root / name for name in _RAPIDOCR_MODEL_NAMES] if model_root else []
    if not model_paths or any(not path.is_file() for path in model_paths):
        _OCR_FINGERPRINT = {
            "engine": engine,
            "engine_version": version,
            "model_fingerprint": "missing",
            "status": "unavailable",
        }
        return dict(_OCR_FINGERPRINT)
    digest = hashlib.sha256()
    for path in model_paths:
        digest.update(path.name.encode("utf-8"))
        digest.update(path.read_bytes())
    _OCR_FINGERPRINT = {
        "engine": engine,
        "engine_version": version,
        "model_fingerprint": digest.hexdigest(),
        "status": "available",
    }
    return dict(_OCR_FINGERPRINT)


@dataclass(frozen=True)
class ParsedItem:
    """解析器输出的单一、仍未嵌入的内容项。"""

    modality: str
    content_kind: str
    raw_text: str = ""
    page_number: int | None = None
    bbox: dict[str, float] | None = None
    asset_bytes: bytes | None = None
    asset_name: str | None = None
    parent_key: str | None = None


@dataclass(frozen=True)
class ParsedDocument:
    """单份资料的解析结果及其可追踪问题。"""

    parser_name: str
    parser_version: str
    items: tuple[ParsedItem, ...] = ()
    issues: tuple[IngestionIssue, ...] = ()
    raw_artifacts: tuple[tuple[str, bytes], ...] = ()


@dataclass(frozen=True)
class PageRouteDecision:
    """页级质量门的固定输出；不会调用模型或依赖运行时随机状态。"""

    route: str
    reason: str
    quality_gate_version: str
    input_summary: dict[str, int]


def inspect_source(path: Path) -> IngestionIssue | None:
    """只检查输入文件的格式、存在性和大小，不执行解析或远程调用。"""
    if not path.is_file():
        return IngestionIssue(code="source_missing", message="资料不存在或不是普通文件", severity=IssueSeverity.ERROR, blocking=True, source_path=str(path))
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        return IngestionIssue(code="unsupported_format", message="首期不支持该资料格式", severity=IssueSeverity.ERROR, blocking=False, source_path=str(path))
    if path.stat().st_size == 0:
        return IngestionIssue(code="empty_source", message="资料为空", severity=IssueSeverity.ERROR, blocking=False, source_path=str(path))
    if not _source_signature_matches(path):
        return IngestionIssue(code="source_format_mismatch", message="文件扩展名与文件签名或容器格式不一致", severity=IssueSeverity.ERROR, blocking=True, source_path=str(path))
    return None


def decide_page_route(path: Path, page_number: int, parsed: ParsedDocument) -> PageRouteDecision:
    """按本地页证据选择唯一处理路线，无法安全判定时明确阻断。"""
    summary = _page_quality_summary(path, page_number, parsed)
    if any(issue.severity is IssueSeverity.ERROR or issue.blocking for issue in parsed.issues):
        return PageRouteDecision("blocked", "parser_error", PAGE_QUALITY_GATE_VERSION, summary)
    if summary["native_text_chars"] < 0:
        return PageRouteDecision("blocked", "native_page_inspection_failed", PAGE_QUALITY_GATE_VERSION, summary)
    if summary["native_text_chars"] == 0 and summary["xobject_count"] == 0 and summary["item_count"] == 0:
        return PageRouteDecision("blank_page", "physical_page_has_no_content", PAGE_QUALITY_GATE_VERSION, summary)
    has_page_render = any(item.content_kind == "page_render" and item.asset_bytes for item in parsed.items)
    if has_page_render and (summary["item_count"] == 1 or summary["coverage_gap"]):
        return PageRouteDecision("remote_page_fallback", "page_render_required_by_quality_gate", PAGE_QUALITY_GATE_VERSION, summary)
    if summary["native_text_chars"] >= 80 and summary["docling_text_chars"] < summary["native_text_chars"] * PAGE_QUALITY_MIN_TEXT_COVERAGE:
        return PageRouteDecision("blocked", "text_coverage_below_fixed_threshold", PAGE_QUALITY_GATE_VERSION, summary)
    if summary["image_items"]:
        return PageRouteDecision("remote_pictures", "picture_items_present", PAGE_QUALITY_GATE_VERSION, summary)
    if summary["item_count"]:
        return PageRouteDecision("local_objects", "structured_local_objects_present", PAGE_QUALITY_GATE_VERSION, summary)
    return PageRouteDecision("blocked", "nonempty_page_has_no_safe_route", PAGE_QUALITY_GATE_VERSION, summary)


def _source_signature_matches(path: Path) -> bool:
    """以文件头和 XLSX 容器条目校验允许来源，不依赖外部 MIME 数据库。"""
    suffix = path.suffix.lower()
    header = path.read_bytes()[:16]
    if suffix in TEXT_SUFFIXES or suffix == ".csv":
        return b"\x00" not in header
    if suffix == ".pdf":
        return header.startswith(b"%PDF-")
    if suffix == ".png":
        return header.startswith(b"\x89PNG\r\n\x1a\n")
    if suffix in {".jpg", ".jpeg"}:
        return header.startswith(b"\xff\xd8\xff")
    if suffix == ".webp":
        return len(header) >= 12 and header[:4] == b"RIFF" and header[8:12] == b"WEBP"
    if suffix in {".tif", ".tiff"}:
        return header.startswith((b"II*\x00", b"MM\x00*"))
    if suffix == ".xlsx":
        if not header.startswith(b"PK\x03\x04"):
            return False
        try:
            import zipfile
            with zipfile.ZipFile(path) as archive:
                return "[Content_Types].xml" in archive.namelist() and any(name.startswith("xl/") for name in archive.namelist())
        except (OSError, zipfile.BadZipFile):
            return False
    return False


def _page_quality_summary(path: Path, page_number: int, parsed: ParsedDocument) -> dict[str, int]:
    """汇总质量门需要的稳定页证据；pypdf 读取失败会以负值使路由阻断。"""
    native_text = ""
    xobjects = 0
    if path.suffix.lower() == ".pdf":
        try:
            from pypdf import PdfReader
            page = PdfReader(path).pages[page_number - 1]
            native_text = (page.extract_text() or "").strip()
            resources = page.get("/Resources") or {}
            xobject_map = resources.get("/XObject") if hasattr(resources, "get") else None
            xobjects = len(xobject_map or {})
        except Exception:
            return {"native_text_chars": -1, "xobject_count": -1, "docling_text_chars": 0, "item_count": len(parsed.items), "image_items": 0, "coverage_gap": 1}
    text_chars = sum(len(item.raw_text.strip()) for item in parsed.items if item.modality in {"text", "table", "equation"})
    image_items = sum(1 for item in parsed.items if item.modality == "image" and item.content_kind != "page_render")
    return {
        "native_text_chars": len(native_text),
        "xobject_count": xobjects,
        "docling_text_chars": text_chars,
        "item_count": len(parsed.items),
        "image_items": image_items,
        "coverage_gap": int(len(native_text) >= 80 and text_chars < len(native_text) * PAGE_QUALITY_MIN_TEXT_COVERAGE),
    }


def parse_document(path: Path, preferred_parser: str) -> ParsedDocument:
    """以 MinerU 优先、Docling 回退的策略解析资料，缺依赖时隔离失败。"""
    suffix = path.suffix.lower()
    if suffix in TEXT_SUFFIXES:
        return _parse_text(path)
    if suffix == ".csv":
        return _parse_csv(path)
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    if suffix in IMAGE_SUFFIXES:
        payload = path.read_bytes()
        return ParsedDocument("image", "builtin-2", (ParsedItem("image", "image", asset_bytes=payload, asset_name=path.name),))
    if suffix in PDF_SUFFIXES:
        return _parse_pdf(path, preferred_parser)
    raise ValueError(f"unsupported source suffix: {suffix}")


def iter_document_pages(path: Path, preferred_parser: str) -> Iterator[ParsedDocument]:
    """逐页产生解析结果；PDF 每页使用独立进程以释放原生模型内存。"""
    if path.suffix.lower() != ".pdf":
        yield parse_document(path, preferred_parser)
        return
    if importlib.util.find_spec("docling") is None:
        yield _parse_pdf(path, preferred_parser)
        return
    yield from _iter_pdf_with_docling(path, preferred_parser)


def parse_document_page(path: Path, preferred_parser: str, page_number: int) -> ParsedDocument:
    """解析指定物理页，供可恢复摄取按页调度。"""
    if page_number < 1:
        raise ValueError("page_number must be positive")
    if path.suffix.lower() != ".pdf":
        if page_number != 1:
            raise ValueError("non-PDF sources contain one logical page")
        return parse_document(path, preferred_parser)
    if importlib.util.find_spec("docling") is None:
        return _parse_pdf(path, preferred_parser)
    return _parse_docling_page_isolated(path, page_number)


def split_text_units(text: str, *, target_chars: int | None = None, max_chars: int | None = None, overlap_chars: int | None = None) -> list[str]:
    """按标题/段落优先、字符上限兜底的确定性切分；参数写入 manifest 的 TEXT_SPLIT。"""
    target = target_chars if target_chars is not None else TEXT_SPLIT["target_chars"]
    maximum = max_chars if max_chars is not None else TEXT_SPLIT["max_chars"]
    overlap = overlap_chars if overlap_chars is not None else TEXT_SPLIT["overlap_chars"]
    if maximum < 1 or target < 1:
        raise ValueError("text split sizes must be positive")
    if overlap < 0 or overlap >= maximum:
        raise ValueError("overlap_chars must be in [0, max_chars)")
    normalized = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not normalized:
        return []
    if len(normalized) <= target:
        return [normalized]
    pieces = _structured_pieces(normalized)
    chunks: list[str] = []
    for piece in pieces:
        if len(piece) <= maximum:
            _append_with_budget(chunks, piece, target=target, maximum=maximum)
        else:
            for hard in _hard_split(piece, maximum=maximum, overlap=overlap):
                _append_with_budget(chunks, hard, target=target, maximum=maximum)
    return chunks or [normalized[:maximum]]


def _structured_pieces(text: str) -> list[str]:
    """先按 ATX 标题，再按空行段落，产出可合并的结构化片段。"""
    lines = text.split("\n")
    sections: list[str] = []
    current: list[str] = []
    for line in lines:
        if line.lstrip().startswith("#") and current:
            sections.append("\n".join(current).strip())
            current = [line]
        else:
            current.append(line)
    if current:
        sections.append("\n".join(current).strip())
    pieces: list[str] = []
    for section in sections:
        paragraphs = [part.strip() for part in section.split("\n\n") if part.strip()]
        pieces.extend(paragraphs or ([section] if section else []))
    return pieces


def _hard_split(text: str, *, maximum: int, overlap: int) -> list[str]:
    """在无结构边界时按字符窗口切分，保留重叠。"""
    if len(text) <= maximum:
        return [text]
    pieces: list[str] = []
    start = 0
    while start < len(text):
        end = min(len(text), start + maximum)
        if end < len(text):
            window = text[start:end]
            break_at = max(window.rfind("\n"), window.rfind("。"), window.rfind(". "), window.rfind("；"), window.rfind("; "))
            if break_at >= maximum // 2:
                end = start + break_at + 1
        pieces.append(text[start:end].strip())
        if end >= len(text):
            break
        start = max(end - overlap, start + 1)
    return [piece for piece in pieces if piece]


def _append_with_budget(chunks: list[str], piece: str, *, target: int, maximum: int) -> None:
    """尽量把相邻段落合并到 target 附近，超过 max 或已达 target 再开新块。"""
    piece = piece.strip()
    if not piece:
        return
    if not chunks:
        chunks.append(piece)
        return
    candidate = f"{chunks[-1]}\n\n{piece}"
    if len(chunks[-1]) < target and len(candidate) <= maximum:
        chunks[-1] = candidate
        return
    chunks.append(piece)


def _parse_text(path: Path) -> ParsedDocument:
    """使用确定性的 UTF-8 文本解析与切分，作为无外部依赖路径。"""
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        text = path.read_text(encoding="utf-8", errors="replace")
    units = split_text_units(text)
    items = tuple(ParsedItem("text", "paragraph", raw_text=chunk) for chunk in units) or (ParsedItem("text", "paragraph", raw_text=""),)
    return ParsedDocument("text", "builtin-1", items)


def _parse_csv(path: Path) -> ParsedDocument:
    """把 CSV 的表头和每一行保留为独立、可检索的表格单元。"""
    try:
        content = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        content = path.read_text(encoding="utf-8", errors="replace")
    rows = list(csv.reader(content.splitlines()))
    if not rows:
        return ParsedDocument("csv", "builtin-1", issues=(_empty_table_issue(path),))
    headers = _table_headers(rows[0])
    items = tuple(
        ParsedItem("table", "table_row", raw_text=_render_table_row(headers, row, row_index))
        for row_index, row in enumerate(rows[1:], 2)
        if any(cell.strip() for cell in row)
    )
    return ParsedDocument("csv", "builtin-1", items or (ParsedItem("table", "table", raw_text="表头：" + "、".join(headers)),))


def _parse_xlsx(path: Path) -> ParsedDocument:
    """使用 openpyxl 的只读模式稳定提取每张工作表的行列语义。"""
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        items: list[ParsedItem] = []
        for sheet in workbook.worksheets:
            rows = [["" if value is None else str(value) for value in row] for row in sheet.iter_rows(values_only=True)]
            if not rows:
                continue
            headers = _table_headers(rows[0])
            for row_index, row in enumerate(rows[1:], 2):
                if any(cell.strip() for cell in row):
                    items.append(ParsedItem("table", "table_row", raw_text=_render_table_row(headers, row, row_index, sheet.title)))
        workbook.close()
    except Exception as exc:
        return ParsedDocument("xlsx", "openpyxl", issues=(IngestionIssue(code="xlsx_parse_failed", message=f"XLSX 解析失败：{type(exc).__name__}", severity=IssueSeverity.ERROR, blocking=False, source_path=str(path)),))
    return ParsedDocument("xlsx", "openpyxl", tuple(items), () if items else (_empty_table_issue(path),))


def _table_headers(row: list[str]) -> list[str]:
    """为缺失表头的列生成稳定名称，避免丢失列语义。"""
    return [cell.strip() or f"列{index}" for index, cell in enumerate(row, 1)]


def _render_table_row(headers: list[str], row: list[str], row_index: int, sheet_name: str | None = None) -> str:
    """按表头、工作表和闭区间行号渲染确定性表格文本。"""
    pairs = [f"{header}：{row[index].strip() if index < len(row) else ''}" for index, header in enumerate(headers)]
    prefix = f"工作表：{sheet_name}\n" if sheet_name else ""
    return prefix + f"行：{row_index}\n" + "\n".join(pairs)


def _empty_table_issue(path: Path) -> IngestionIssue:
    """为无可索引行的表格资料创建可追踪问题。"""
    return IngestionIssue(code="table_empty", message="表格没有可索引的数据行", severity=IssueSeverity.WARNING, blocking=False, source_path=str(path))


def _parse_pdf(path: Path, preferred_parser: str) -> ParsedDocument:
    """选择已安装的指定 PDF adapter；不以 PyMuPDF 作为新链路回退。"""
    if importlib.util.find_spec("docling") is None:
        return ParsedDocument(
            "unavailable", "0",
            issues=(IngestionIssue(code="parser_dependency_missing", message="PDF 解析依赖未安装；资料已隔离，未调用视觉服务", severity=IssueSeverity.ERROR, blocking=False, source_path=str(path)),),
        )
    return _parse_pdf_with_docling(path, preferred_parser)


def _parse_pdf_with_docling(path: Path, preferred_parser: str) -> ParsedDocument:
    """聚合逐页结果，供兼容调用者和小型测试使用。"""
    pages = list(_iter_pdf_with_docling(path, preferred_parser))
    items = [item for page in pages for item in page.items]
    issues = [issue for page in pages for issue in page.issues]
    raw_artifacts = [artifact for page in pages for artifact in page.raw_artifacts]
    if not items:
        return ParsedDocument(
            "docling", "2.115.0",
            issues=(IngestionIssue(code="pdf_empty_output", message="Docling 未提取到可索引正文", severity=IssueSeverity.ERROR, blocking=False, source_path=str(path)),),
        )
    if preferred_parser == "mineru":
        issues.append(IngestionIssue(code="mineru_fallback_docling", message="MinerU 未配置为可运行解析路径，已使用 Docling fallback", severity=IssueSeverity.WARNING, blocking=False, source_path=str(path)))
    return ParsedDocument("docling", "2.115.0", tuple(items), tuple(issues), tuple(raw_artifacts))


def _iter_pdf_with_docling(path: Path, preferred_parser: str) -> Iterator[ParsedDocument]:
    """按物理页隔离运行 Docling，并为每页保留成功或失败证据。"""
    from pypdf import PdfReader

    page_count = len(PdfReader(path).pages)
    for page_number in range(1, page_count + 1):
        yield _parse_docling_page_isolated(path, page_number)


def _parse_docling_page_isolated(path: Path, page_number: int) -> ParsedDocument:
    """在 spawn 子进程中解析一页，超时或崩溃时返回阻断 issue。"""
    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    process = context.Process(target=_docling_page_worker, args=(str(path), page_number, sender))
    process.start()
    sender.close()
    timeout = int(os.getenv("MULTIMODAL_DOCLING_PAGE_TIMEOUT_SECONDS", "900"))
    try:
        if receiver.poll(timeout):
            parsed = receiver.recv()
            process.join(30)
            return parsed
        process.terminate()
        process.join(30)
        return _page_failure(path, page_number, "TimeoutError")
    except (EOFError, OSError):
        process.join(30)
        return _page_failure(path, page_number, f"ProcessExit{process.exitcode}")
    finally:
        receiver.close()


def _docling_page_worker(path_value: str, page_number: int, sender: object) -> None:
    """子进程入口：解析单页并通过管道返回纯本地结果。"""
    path = Path(path_value)
    try:
        from docling.datamodel.pipeline_options import PdfPipelineOptions

        artifacts_path = Path(os.getenv("MULTIMODAL_DOCLING_ARTIFACTS_DIR", Path.home() / ".cache" / "docling" / "models"))
        options = PdfPipelineOptions(artifacts_path=artifacts_path)
        options.do_ocr = False
        options.generate_picture_images = True
        options.generate_page_images = True
        converter = _new_docling_converter(options)
        result, _ = _convert_docling_page(path, page_number, options, converter)
        items_list, issues = _docling_page_items(result.document, path, page_number)
        artifact = (
            f"docling_page_{page_number:04d}.json",
            json.dumps(result.document.export_to_dict(), ensure_ascii=False, sort_keys=True).encode("utf-8"),
        )
        sender.send(ParsedDocument("docling", "2.115.0", tuple(items_list), issues, (artifact,)))
    except Exception as exc:
        sender.send(_page_failure(path, page_number, type(exc).__name__))
    finally:
        sender.close()


def _page_failure(path: Path, page_number: int, error_name: str) -> ParsedDocument:
    """为单页隔离失败创建可审计的阻断结果。"""
    issue = IngestionIssue(
        code="pdf_page_parse_failed",
        message=f"Docling 第 {page_number} 页解析失败：{error_name}",
        severity=IssueSeverity.ERROR,
        blocking=True,
        source_path=str(path),
    )
    return ParsedDocument("docling", "2.115.0", issues=(issue,))


def _pdf_page_has_content(path: Path, page_number: int) -> bool:
    """独立检查 PDF 页是否包含文本或图片对象，防止把解析遗漏当作空白页。"""
    from pypdf import PdfReader

    page = PdfReader(path).pages[page_number - 1]
    if (page.extract_text() or "").strip():
        return True
    resources = page.get("/Resources") or {}
    xobjects = resources.get("/XObject") if hasattr(resources, "get") else None
    return bool(xobjects)


def _new_docling_converter(options: object) -> object:
    """创建使用同一离线模型配置的 Docling converter。"""
    from docling.datamodel.base_models import InputFormat
    from docling.document_converter import DocumentConverter, PdfFormatOption

    return DocumentConverter(format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=options)})


def _convert_docling_page(path: Path, page_number: int, options: object, converter: object) -> tuple[object, object]:
    """转换单页；converter 状态异常时仅重建并重试一次。"""
    try:
        return converter.convert(path, page_range=(page_number, page_number)), converter
    except Exception:
        replacement = _new_docling_converter(options)
        return replacement.convert(path, page_range=(page_number, page_number)), replacement


def _docling_items(document: object, source_path: Path) -> tuple[list[ParsedItem], list[IngestionIssue]]:
    """将 Docling 项映射为项目契约；图片只保留本地 caption 与原始资源。"""
    from docling_core.types.doc import FormulaItem, PictureItem, TableItem, TextItem

    items: list[ParsedItem] = []
    ocr_issues: list[IngestionIssue] = []
    for index, (item, _) in enumerate(document.iterate_items(), 1):
        page_number, bbox = _docling_locator(item, document)
        if isinstance(item, TableItem):
            text = item.export_to_markdown(doc=document).strip()
            if text:
                items.append(ParsedItem("table", "table", raw_text=text, page_number=page_number, bbox=bbox, parent_key=f"page_{page_number}" if page_number else None))
        elif isinstance(item, FormulaItem):
            formula = (item.text or "").strip()
            if formula:
                items.append(ParsedItem("equation", "formula", raw_text=formula, page_number=page_number, bbox=bbox, parent_key=f"page_{page_number}" if page_number else None))
        elif isinstance(item, PictureItem):
            asset_bytes = _docling_image_bytes(item, document)
            if asset_bytes:
                caption = (item.caption_text(document) or "").strip()
                items.append(ParsedItem("image", "image", raw_text=caption, page_number=page_number, bbox=bbox, asset_bytes=asset_bytes, asset_name=f"page_{page_number or 0:04d}_image_{index:04d}.png", parent_key=f"page_{page_number}" if page_number else None))
            else:
                ocr_issues.append(IngestionIssue(code="image_asset_missing", message=f"第 {page_number or 0} 页图片资源无法提取", severity=IssueSeverity.WARNING, blocking=False, source_path=str(source_path)))
        elif isinstance(item, TextItem) and (item.text or "").strip():
            label = str(getattr(getattr(item, "label", None), "value", getattr(item, "label", "paragraph")))
            items.append(ParsedItem("text", label or "paragraph", raw_text=item.text.strip(), page_number=page_number, bbox=bbox, parent_key=f"page_{page_number}" if page_number else None))
    return items, ocr_issues


def _docling_page_items(document: object, source_path: Path, page_number: int) -> tuple[list[ParsedItem], tuple[IngestionIssue, ...]]:
    """将无结构化项但含 PDF 内容的页作为 Docling 整页图像保留给远程图片链。"""
    items, issues = _docling_items(document, source_path)
    if items:
        return items, tuple(issues)
    if not _pdf_page_has_content(source_path, page_number):
        return items, tuple(issues) + (IngestionIssue(
            code="pdf_page_empty", message=f"Docling 第 {page_number} 页未返回可索引内容",
            severity=IssueSeverity.WARNING, blocking=False, source_path=str(source_path),
        ),)
    page_image = _docling_page_image_bytes(document, page_number)
    if page_image is not None:
        return [ParsedItem(
            "image", "page_render", page_number=page_number, asset_bytes=page_image,
            asset_name=f"page_{page_number:04d}_render.png", parent_key=f"page_{page_number}",
        )], tuple(issues)
    return items, tuple(issues) + (IngestionIssue(
        code="pdf_page_content_missing", message=f"Docling 第 {page_number} 页未返回可索引内容",
        severity=IssueSeverity.ERROR, blocking=True, source_path=str(source_path),
    ),)


def _docling_locator(item: object, document: object) -> tuple[int | None, dict[str, float] | None]:
    """把 Docling 的页码和坐标转换为一基、左上原点的归一化 bbox。"""
    provenance = next(iter(getattr(item, "prov", []) or []), None)
    if provenance is None:
        return None, None
    page_number = int(provenance.page_no)
    bbox = getattr(provenance, "bbox", None)
    page = getattr(document, "pages", {}).get(page_number)
    size = getattr(page, "size", None)
    if bbox is None or size is None or not size.width or not size.height:
        return page_number, None
    x0, x1 = sorted((bbox.l / size.width, bbox.r / size.width))
    y0, y1 = sorted((bbox.t / size.height, bbox.b / size.height))
    origin = str(getattr(bbox.coord_origin, "value", bbox.coord_origin)).lower()
    if origin.startswith("bottom"):
        y0, y1 = 1 - y1, 1 - y0
    if x0 >= x1 or y0 >= y1 or min(x0, y0) < 0 or max(x1, y1) > 1:
        return page_number, None
    return page_number, {"x0": x0, "y0": y0, "x1": x1, "y1": y1}


def _docling_image_bytes(item: object, document: object) -> bytes | None:
    """把可用图片渲染为 PNG；没有可靠图片时返回空而非猜测资源。"""
    try:
        image = item.get_image(document)
        if image is None:
            return None
        output = BytesIO()
        image.save(output, format="PNG")
        return output.getvalue()
    except Exception:
        return None


def _docling_page_image_bytes(document: object, page_number: int) -> bytes | None:
    """导出 Docling 已生成的整页图像，不引入额外 PDF 解析器。"""
    try:
        page = getattr(document, "pages", {}).get(page_number)
        image = getattr(getattr(page, "image", None), "pil_image", None)
        if image is None:
            return None
        output = BytesIO()
        image.save(output, format="PNG")
        return output.getvalue()
    except Exception:
        return None


def _rapidocr_probe(image_bytes: bytes) -> OcrResult:
    """使用本地 RapidOCR 为图片生成分类 OCR 结果，永不抛异常。

    ponytail: 把原来吞所有异常返回空串的 helper 拆成可观测契约；
    状态机 ok|no_text|failed|engine_unavailable 覆盖 §5.4 全部分类。
    """
    started = time.monotonic()
    fp = ocr_fingerprint()
    if fp["status"] != "available":
        return OcrResult("", fp["engine"], fp["engine_version"], fp["model_fingerprint"], 0, 0.0, 0, "engine_unavailable")
    global _RAPID_OCR_ENGINE
    if _RAPID_OCR_ENGINE is None:
        try:
            from rapidocr import RapidOCR
            _RAPID_OCR_ENGINE = RapidOCR()
        except Exception as exc:
            elapsed_ms = int((time.monotonic() - started) * 1000)
            return OcrResult("", fp["engine"], fp["engine_version"], fp["model_fingerprint"], 0, 0.0, elapsed_ms, "engine_unavailable", type(exc).__name__)
    try:
        result = _RAPID_OCR_ENGINE(image_bytes)
        texts = [text.strip() for text in (result.txts or ()) if text and text.strip()]
        scores = [float(score) for score in (getattr(result, "scores", None) or ()) if score is not None]
        mean_confidence = round(sum(scores) / len(scores), 6) if scores else 0.0
        elapsed_ms = int((time.monotonic() - started) * 1000)
        if not texts:
            return OcrResult("", fp["engine"], fp["engine_version"], fp["model_fingerprint"], 0, 0.0, elapsed_ms, "no_text")
        return OcrResult("\n".join(texts), fp["engine"], fp["engine_version"], fp["model_fingerprint"], len(texts), mean_confidence, elapsed_ms, "ok")
    except Exception as exc:
        elapsed_ms = int((time.monotonic() - started) * 1000)
        return OcrResult("", fp["engine"], fp["engine_version"], fp["model_fingerprint"], 0, 0.0, elapsed_ms, "failed", type(exc).__name__)


def _ocr_issues(probe: OcrResult, path: Path) -> tuple[IngestionIssue, ...]:
    """把 OCR 分类状态映射为可追踪 issue，不记录图片正文。"""
    if probe.status == "engine_unavailable":
        return (IngestionIssue(code="ocr_engine_unavailable", message="RapidOCR 引擎或模型不可用，无法识别图片文字", severity=IssueSeverity.ERROR, blocking=True, source_path=str(path)),)
    if probe.status == "failed":
        return (IngestionIssue(code="ocr_failed", message=f"RapidOCR 推理失败：{probe.failure_type or 'UnknownError'}，已隔离该图片", severity=IssueSeverity.WARNING, blocking=False, source_path=str(path)),)
    if probe.status == "no_text":
        return (IngestionIssue(code="ocr_no_text", message="图片未识别到文字，可由后续 VLM 判断是否为信息图", severity=IssueSeverity.WARNING, blocking=False, source_path=str(path)),)
    return ()
